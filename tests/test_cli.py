from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import openpyxl
import pandas as pd


REQUIRED_COLUMNS = [
    "nickname",
    "profile_url",
    "fans",
    "category",
    "sales_30d",
    "gmv_30d",
    "product_count",
    "live_count_7d",
    "median_likes",
]


def make_csv(path: Path, rows: list[dict]) -> None:
    pd.DataFrame(rows, columns=REQUIRED_COLUMNS).to_csv(path, index=False)


def run_cli(*args: str) -> dict | list:
    completed = subprocess.run(
        [sys.executable, "-m", "app.cli", *args],
        check=True,
        capture_output=True,
        text=True,
    )
    return json.loads(completed.stdout)


def test_cli_import_updates_profile_url_and_lists_records(tmp_path: Path) -> None:
    db_path = tmp_path / "creators.db"
    csv1 = tmp_path / "batch1.csv"
    csv2 = tmp_path / "batch2.csv"

    make_csv(
        csv1,
        [
            {
                "nickname": "A",
                "profile_url": "https://example.com/a",
                "fans": 100,
                "category": "pet",
                "sales_30d": 50,
                "gmv_30d": 500,
                "product_count": 10,
                "live_count_7d": 2,
                "median_likes": 20,
            },
            {
                "nickname": "B",
                "profile_url": "https://example.com/b",
                "fans": 200,
                "category": "pet",
                "sales_30d": 80,
                "gmv_30d": 1000,
                "product_count": 8,
                "live_count_7d": 3,
                "median_likes": 30,
            },
        ],
    )
    make_csv(
        csv2,
        [
            {
                "nickname": "A-updated",
                "profile_url": "https://example.com/a",
                "fans": 100,
                "category": "pet-food",
                "sales_30d": 90,
                "gmv_30d": 900,
                "product_count": 12,
                "live_count_7d": 4,
                "median_likes": 25,
            }
        ],
    )

    first_import = run_cli("import-csv", "--db", str(db_path), "--csv", str(csv1))
    second_import = run_cli("import-csv", "--db", str(db_path), "--csv", str(csv2))
    records = run_cli("list", "--db", str(db_path))

    assert first_import["imported"] == 2
    assert len(first_import["records"]) == 2
    assert second_import["imported"] == 1
    assert len(second_import["records"]) == 2
    assert isinstance(records, list)
    assert len(records) == 2

    row_a = next(row for row in records if row["profile_url"] == "https://example.com/a")
    assert row_a["nickname"] == "A-updated"
    assert row_a["category"] == "pet-food"
    assert row_a["sales_30d"] == 90
    assert row_a["sales_per_fan"] == 0.9
    assert row_a["score"] > 0


def test_cli_export_writes_excel_file(tmp_path: Path) -> None:
    db_path = tmp_path / "creators.db"
    csv_path = tmp_path / "batch.csv"
    export_path = tmp_path / "ranking.xlsx"

    make_csv(
        csv_path,
        [
            {
                "nickname": "A",
                "profile_url": "https://example.com/a",
                "fans": 100,
                "category": "pet",
                "sales_30d": 50,
                "gmv_30d": 500,
                "product_count": 10,
                "live_count_7d": 2,
                "median_likes": 20,
            }
        ],
    )

    run_cli("import-csv", "--db", str(db_path), "--csv", str(csv_path))
    result = run_cli("export-excel", "--db", str(db_path), "--output", str(export_path))

    workbook = openpyxl.load_workbook(export_path)
    sheet = workbook["creators"]

    assert result["output_path"] == str(export_path)
    assert export_path.exists()
    assert sheet.max_row == 2
    assert sheet.max_column >= 11


def test_cli_export_respects_filters(tmp_path: Path) -> None:
    db_path = tmp_path / "creators.db"
    csv_path = tmp_path / "batch.csv"
    export_path = tmp_path / "filtered.xlsx"

    make_csv(
        csv_path,
        [
            {
                "nickname": "A",
                "profile_url": "https://example.com/a",
                "fans": 100,
                "category": "pet",
                "sales_30d": 50,
                "gmv_30d": 500,
                "product_count": 10,
                "live_count_7d": 2,
                "median_likes": 20,
            },
            {
                "nickname": "B",
                "profile_url": "https://example.com/b",
                "fans": 500,
                "category": "pet-food",
                "sales_30d": 300,
                "gmv_30d": 3000,
                "product_count": 30,
                "live_count_7d": 8,
                "median_likes": 80,
            },
        ],
    )

    run_cli("import-csv", "--db", str(db_path), "--csv", str(csv_path))
    run_cli(
        "export-excel",
        "--db",
        str(db_path),
        "--output",
        str(export_path),
        "--category",
        "pet-food",
        "--fans-min",
        "200",
        "--fans-max",
        "1000",
    )

    workbook = openpyxl.load_workbook(export_path)
    sheet = workbook["creators"]

    assert sheet.max_row == 2
    assert sheet["A2"].value == "B"


def test_cli_export_filters_dadaduo_specific_fields(tmp_path: Path) -> None:
    db_path = tmp_path / "creators.db"
    xlsx_path = tmp_path / "dadaduo.xlsx"
    export_path = tmp_path / "filtered.xlsx"

    pd.DataFrame(
        [
            {
                "达人名称": "A",
                "达人抖音id": "a",
                "带货等级": 2,
                "视频数": 10,
                "粉丝总量": 1000,
                "平均获赞数": 10,
                "平均赞粉比": "1.00%",
                "推广商品数": 1,
                "预估结算金额": "0~250",
            },
            {
                "达人名称": "B",
                "达人抖音id": "b",
                "带货等级": 5,
                "视频数": 120,
                "粉丝总量": 9000,
                "平均获赞数": 600,
                "平均赞粉比": "6.67%",
                "推广商品数": 30,
                "预估结算金额": "10w~50w",
            },
        ]
    ).to_excel(xlsx_path, index=False)

    run_cli("import-csv", "--db", str(db_path), "--csv", str(xlsx_path))
    run_cli(
        "export-excel",
        "--db",
        str(db_path),
        "--output",
        str(export_path),
        "--creator-level-min",
        "4",
        "--settlement-min",
        "100000",
        "--product-count-min",
        "10",
    )

    workbook = openpyxl.load_workbook(export_path)
    sheet = workbook["creators"]

    assert sheet.max_row == 2
    assert sheet["A2"].value == "B"
